"""英语单词 API 路由"""
import io
import csv
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from sqlalchemy import or_

from ..database import get_db
from ..models.word import Word, WordBook
from ..schemas.word import (
    WordCreate, WordUpdate, WordOut, WordBookOut, WordImportResult,
)

router = APIRouter()


# ─── 词库管理 ───────────────────────────────────────────────

@router.get("/books", response_model=List[WordBookOut], summary="获取所有词库")
def list_books(db: Session = Depends(get_db)):
    return db.query(WordBook).order_by(WordBook.grade, WordBook.semester).all()


@router.post("/books", response_model=WordBookOut, summary="创建词库")
def create_book(
    name: str = Query(...),
    grade: int = Query(..., ge=1, le=6),
    semester: str = Query("上"),
    publisher: str = Query("人教版PEP"),
    db: Session = Depends(get_db),
):
    book = WordBook(name=name, grade=grade, semester=semester, publisher=publisher)
    db.add(book)
    db.commit()
    db.refresh(book)
    return book


# ─── 单词 CRUD ──────────────────────────────────────────────

@router.get("/", response_model=List[WordOut], summary="查询单词")
def list_words(
    book_id: Optional[int] = Query(None),
    grade: Optional[int] = Query(None, ge=1, le=6),
    keyword: Optional[str] = Query(None, description="搜索单词或释义"),
    difficulty: Optional[int] = Query(None, ge=1, le=5),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    q = db.query(Word)
    if book_id:
        q = q.filter(Word.book_id == book_id)
    if grade:
        q = q.join(WordBook).filter(WordBook.grade == grade)
    if keyword:
        q = q.filter(or_(Word.word.contains(keyword), Word.meaning.contains(keyword)))
    if difficulty:
        q = q.filter(Word.difficulty == difficulty)
    return q.offset((page - 1) * page_size).limit(page_size).all()


@router.get("/count", summary="单词总数统计")
def word_count(
    book_id: Optional[int] = Query(None),
    grade: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(Word)
    if book_id:
        q = q.filter(Word.book_id == book_id)
    if grade:
        q = q.join(WordBook).filter(WordBook.grade == grade)
    return {"count": q.count()}


@router.post("/", response_model=WordOut, summary="添加单词")
def create_word(
    book_id: int = Query(...),
    data: WordCreate = ...,
    db: Session = Depends(get_db),
):
    book = db.query(WordBook).get(book_id)
    if not book:
        raise HTTPException(404, "词库不存在")
    existing = db.query(Word).filter(Word.book_id == book_id, Word.word == data.word).first()
    if existing:
        raise HTTPException(409, f"单词 '{data.word}' 已存在于该词库")
    word = Word(book_id=book_id, **data.model_dump())
    db.add(word)
    book.word_count = db.query(Word).filter(Word.book_id == book_id).count() + 1
    db.commit()
    db.refresh(word)
    return word


@router.put("/{word_id}", response_model=WordOut, summary="更新单词")
def update_word(word_id: int, data: WordUpdate, db: Session = Depends(get_db)):
    word = db.query(Word).get(word_id)
    if not word:
        raise HTTPException(404, "单词不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(word, k, v)
    db.commit()
    db.refresh(word)
    return word


@router.delete("/{word_id}", summary="删除单词")
def delete_word(word_id: int, db: Session = Depends(get_db)):
    word = db.query(Word).get(word_id)
    if not word:
        raise HTTPException(404, "单词不存在")
    book_id = word.book_id
    db.delete(word)
    db.commit()
    book = db.query(WordBook).get(book_id)
    if book:
        book.word_count = db.query(Word).filter(Word.book_id == book_id).count()
        db.commit()
    return {"message": "已删除"}


# ─── 批量导入 ───────────────────────────────────────────────

@router.post("/import", response_model=WordImportResult, summary="导入单词（CSV/Excel）")
async def import_words(
    file: UploadFile = File(...),
    book_id: int = Query(..., description="目标词库ID"),
    db: Session = Depends(get_db),
):
    """
    支持 CSV 和 Excel(.xlsx) 格式导入。
    CSV 列顺序：word, phonetic, pos, meaning, unit, difficulty, tags
    Excel 列名：word/单词, phonetic/音标, pos/词性, meaning/释义, unit/单元, difficulty/难度, tags/标签
    """
    book = db.query(WordBook).get(book_id)
    if not book:
        raise HTTPException(404, "词库不存在")

    content = await file.read()
    filename = file.filename or ""

    rows = []
    errors = []

    if filename.endswith(".csv"):
        rows, errors = _parse_csv(content)
    elif filename.endswith((".xlsx", ".xls")):
        rows, errors = _parse_excel(content)
    else:
        raise HTTPException(400, "仅支持 .csv 和 .xlsx/.xls 格式")

    imported = 0
    skipped = 0
    existing_words = {
        w.word.lower()
        for w in db.query(Word.word).filter(Word.book_id == book_id).all()
    }

    for i, row in enumerate(rows, start=2):  # 第2行开始（第1行是表头）
        word_text = row.get("word", "").strip()
        meaning = row.get("meaning", "").strip()
        if not word_text or not meaning:
            errors.append(f"第{i}行：缺少word或meaning字段")
            continue
        if word_text.lower() in existing_words:
            skipped += 1
            continue
        try:
            w = Word(
                book_id=book_id,
                word=word_text,
                phonetic=row.get("phonetic", "").strip(),
                pos=row.get("pos", "").strip(),
                meaning=meaning,
                unit=row.get("unit", "").strip(),
                difficulty=int(row.get("difficulty", 1) or 1),
                tags=row.get("tags", "").strip(),
            )
            db.add(w)
            existing_words.add(word_text.lower())
            imported += 1
        except Exception as e:
            errors.append(f"第{i}行：{str(e)}")

    book.word_count = db.query(Word).filter(Word.book_id == book_id).count() + imported
    db.commit()

    return WordImportResult(
        total=len(rows), imported=imported, skipped=skipped, errors=errors[:50]
    )


def _parse_csv(content: bytes):
    """解析CSV，支持UTF-8和GBK编码"""
    rows, errors = [], []
    for encoding in ("utf-8-sig", "gbk", "utf-8"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        return [], ["无法识别文件编码，请使用UTF-8或GBK编码"]

    reader = csv.DictReader(io.StringIO(text))
    # 兼容中文列名
    field_map = {
        "单词": "word", "英文": "word",
        "音标": "phonetic",
        "词性": "pos",
        "释义": "meaning", "中文": "meaning", "中文释义": "meaning",
        "单元": "unit",
        "难度": "difficulty",
        "标签": "tags",
    }
    for row in reader:
        normalized = {}
        for k, v in row.items():
            key = field_map.get(k.strip(), k.strip().lower())
            normalized[key] = v or ""
        rows.append(normalized)
    return rows, errors


def _parse_excel(content: bytes):
    """解析Excel文件"""
    try:
        import openpyxl
    except ImportError:
        return [], ["服务器缺少openpyxl库，无法解析Excel文件"]

    rows, errors = [], []
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    field_map = {
        "单词": "word", "英文": "word", "word": "word",
        "音标": "phonetic", "phonetic": "phonetic",
        "词性": "pos", "pos": "pos",
        "释义": "meaning", "中文": "meaning", "中文释义": "meaning", "meaning": "meaning",
        "单元": "unit", "unit": "unit",
        "难度": "difficulty", "difficulty": "difficulty",
        "标签": "tags", "tags": "tags",
    }
    mapped_headers = [field_map.get(h, h.lower()) for h in headers]

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(mapped_headers):
                row_dict[mapped_headers[i]] = str(val) if val is not None else ""
        rows.append(row_dict)

    wb.close()
    return rows, errors
